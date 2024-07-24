import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def create_hyperlinked_excel(csv_files, text_cols, link_cols, output_excel):
    # Check if the Excel file already exists
    if os.path.exists(output_excel):
        # Load the existing workbook
        wb = load_workbook(output_excel)
    else:
        # Initialize a new workbook
        wb = Workbook()

    # Iterate over each CSV file and create a new sheet for each
    for csv_file in csv_files:
        # Read the CSV file
        df = pd.read_csv(csv_file)
        
        # Extract the base name of the CSV file to name the sheet
        sheet_name = csv_file.split('/')[-1].split('.')[0]

        # Check if the sheet already exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Create a new sheet with the name based on the CSV file name
            ws = wb.create_sheet(title=sheet_name)

        # Add headers if the sheet is newly created
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(['Website URL'] + text_cols)

        # Iterate over the rows of the DataFrame
        for idx, row in df.iterrows():
            # Get the website URL from the CSV
            website_url = row['Website']
            
            # Create a list for the new row with the URL first
            new_row = [website_url]

            # Iterate over the text and link columns to create hyperlinked text
            for text_col, link_col in zip(text_cols, link_cols):
                cell_text = row[text_col]
                cell_link = row[link_col]
                hyperlinked_text = f'=HYPERLINK("{cell_link}", "{cell_text}")'
                new_row.append(hyperlinked_text)

            # Append the new row to the worksheet
            ws.append(new_row)

    # Save the workbook to the specified output file
    wb.save(output_excel)

# Usage example
csv_files = ['./csvs/nema.csv']
text_cols = ['Quick Change Connectors or Disconnect Connectors or Change Connectors Found in site? (YES/NO)', 'Fluid Connectors or Hydraulic Connectors or Fluid or Hydraulic Found in site? (YES/NO)', 'Coupling (YES/NO)']
link_cols = ['Quick Change Connector Links', 'Fluid Connectors Links', 'Coupling Links']
output_excel = 'output.xlsx'

create_hyperlinked_excel(csv_files, text_cols, link_cols, output_excel)
